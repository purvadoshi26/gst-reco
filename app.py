"""
GST Audit Reconciliation Tool — Expert Edition
Built from real audit data analysis (Akin Chemicals, Vizag, Jan 2026)
Handles: ITC Reco (GSTR-2B vs Tally) + Sales Reco (GSTR-1 vs Books)
"""
import streamlit as st
import io, re, datetime
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="GST Audit Reco Tool", page_icon="📊",
                   layout="wide", initial_sidebar_state="expanded")
st.markdown("""
<style>
.hdr{background:#1F3864;padding:1.1rem 1.5rem;border-radius:8px;margin-bottom:1.2rem}
.hdr h1{color:#fff;margin:0;font-size:1.5rem}
.hdr p{color:#AACCEE;margin:3px 0 0;font-size:.88rem}
.fmtbox{background:#EEF2F8;border-left:4px solid #1F3864;padding:.7rem 1rem;
        border-radius:0 6px 6px 0;font-size:.83rem;margin:.4rem 0 .8rem}
div[data-testid="stDownloadButton"] button{
    background:#1F3864;color:white;border:none;padding:.6rem 1.2rem;
    border-radius:6px;font-size:1rem;font-weight:600;width:100%}
</style>""", unsafe_allow_html=True)

# ── Excel styles ──────────────────────────────────────────────────────────────
NAVY="1F3864";BLUE="2E75B6";GREEN="375623";TITLE="203864"
MATCH="D6E4BC";MIS="FFD7D7";ONLY2="FFD966";ONLY1="BDD7EE";RCM_C="E2EFDA";CUST_C="FCE4D6";EXEMPT="F0F0F0"
INR="#,##0.00";NUM="#,##0"
def F(h):  return PatternFill("solid",start_color=h,end_color=h)
def HF(sz=9,c="FFFFFF"): return Font(name="Arial",bold=True,size=sz,color=c)
def DF(sz=9,c="000000",bold=False): return Font(name="Arial",size=sz,color=c,bold=bold)
def BD():
    s=Side(style="thin",color="CCCCCC"); return Border(left=s,right=s,top=s,bottom=s)
def CTR(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
def LFT(): return Alignment(horizontal="left",vertical="center")
def RGT(): return Alignment(horizontal="right",vertical="center")
def safe_float(v):
    try: return float(str(v or "0").replace(",",""))
    except: return 0.0
def clean_gstin(s): return re.sub(r"[^A-Z0-9]","",str(s or "").upper().strip())

# ═══════════════════════════════════════════════════════════════════════════════
#  ITC RECO
# ═══════════════════════════════════════════════════════════════════════════════
def parse_purchase_register(fb):
    wb  = openpyxl.load_workbook(io.BytesIO(fb), data_only=True)
    ws  = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # Find header row
    hr = None
    for i,row in enumerate(all_rows[:20],1):
        v=[str(v or "").strip().lower() for v in row]
        if "date" in v and "particulars" in v: hr=i; break
    if not hr: raise ValueError("Header row not found. Need 'Date' and 'Particulars' columns.")

    hdrs=[str(ws.cell(hr,c).value or "").strip() for c in range(1,ws.max_column+1)]
    hl  =[h.lower() for h in hdrs]

    # Detect if FORMAT A (columnar — preferred, detailed register)
    # A detailed columnar register has: GSTIN column + tax columns with actual numeric values
    igst_cols=[i for i,h in enumerate(hl) if "igst" in h and "input" in h and "custom" not in h]
    cgst_cols=[i for i,h in enumerate(hl) if "cgst" in h and "input" in h]
    sgst_cols=[i for i,h in enumerate(hl) if "sgst" in h and "input" in h]
    cust_cols=[i for i,h in enumerate(hl) if ("custom duty" in h or "customs" in h) and "input" not in h]
    cust_igst=[i for i,h in enumerate(hl) if "igst" in h and "custom" in h]
    has_gstin= any("gstin" in h for h in hl)

    format_a = False
    if has_gstin and (igst_cols or cgst_cols or sgst_cols):
        # Check across ALL data rows (not just first 5) — first rows may be storage-only (CGST/SGST)
        all_tax_cols = igst_cols + cgst_cols + sgst_cols
        check_rows   = all_rows[hr:hr+min(100, len(all_rows)-hr)]
        for r in check_rows:
            for c in all_tax_cols:
                if c < len(r) and isinstance(r[c], (int,float)) and r[c] > 0:
                    format_a = True; break
            if format_a: break

    rows=[]
    if format_a:
        ci={}
        for i,h in enumerate(hl):
            if h=="date"                              and "date"  not in ci: ci["date"] =i
            if "particulars" in h                     and "party" not in ci: ci["party"]=i
            if "voucher no"  in h                     and "vch"   not in ci: ci["vch"]  =i
            if "supplier invoice no" in h             and "sinv"  not in ci: ci["sinv"] =i
            if "supplier invoice date" in h           and "sdt"   not in ci: ci["sdt"]  =i
            if "gstin" in h                           and "gstin" not in ci: ci["gstin"]=i
            if h=="value"                             and "val"   not in ci: ci["val"]  =i

        for row in all_rows[hr:]:
            if not any(row): continue
            dt=row[ci.get("date",0)]
            if not isinstance(dt,(datetime.datetime,datetime.date)): continue
            p=str(row[ci.get("party",1)] or "").strip()
            if not p or p.lower() in ("grand total","total",""): continue
            vch   =str(row[ci.get("vch",2)]   or "").strip()
            sinv  =str(row[ci.get("sinv",3)]  or "").strip()
            sdt   =row[ci.get("sdt",4)]
            gstin =clean_gstin(row[ci.get("gstin",5)])
            val   =safe_float(row[ci.get("val",13)])
            igst  =sum(safe_float(row[c]) for c in igst_cols  if c<len(row))
            cgst  =sum(safe_float(row[c]) for c in cgst_cols  if c<len(row))
            sgst  =sum(safe_float(row[c]) for c in sgst_cols  if c<len(row))
            c_duty=sum(safe_float(row[c]) for c in cust_igst  if c<len(row))
            cat   ="Custom Duty / Import" if c_duty>0 else ("No ITC / RCM" if igst==cgst==sgst==0 else "Regular B2B")
            rows.append({"Date":dt,"Party":p,"GSTIN":gstin,"VchNo":vch,"SupplierInv":sinv,
                         "SupplierDate":sdt,"Value":val,"IGST":igst,"CGST":cgst,"SGST":sgst,
                         "TotalTax":igst+cgst+sgst,"Category":cat})
    else:
        # Format B: ledger sub-row
        cur={}; ig=cg=sg=tx=0.0
        def save():
            if cur.get("p") and cur.get("v"):
                cat="Custom Duty / Import" if cur.get("cust",0)>0 else ("No ITC / RCM" if ig==cg==sg==0 else "Regular B2B")
                rows.append({"Date":cur["d"],"Party":cur["p"],"GSTIN":cur.get("g",""),
                              "VchNo":cur["v"],"SupplierInv":cur.get("si",""),"SupplierDate":None,
                              "Value":tx,"IGST":ig,"CGST":cg,"SGST":sg,"TotalTax":ig+cg+sg,"Category":cat})
        for row in all_rows[hr:]:
            if not any(row): continue
            r=[row[i] if i<len(row) else None for i in range(8)]
            dt,part,vn,si,_,_,db,cr=r
            if isinstance(dt,(datetime.datetime,datetime.date)):
                save(); p=str(part or "").strip()
                if not p or p.lower() in ("total","grand total"): cur={}; ig=cg=sg=tx=0.0; continue
                cur={"d":dt,"p":p,"v":str(vn or ""),"si":str(si or ""),"g":""}; ig=cg=sg=tx=0.0
            elif part and cur:
                lbl=str(part).upper(); amt=safe_float(db)
                if "IGST" in lbl and "CUSTOM" in lbl: cur["cust"]=cur.get("cust",0)+amt; ig+=amt
                elif "IGST" in lbl: ig+=amt
                elif "CGST" in lbl: cg+=amt
                elif "SGST" in lbl or "UTGST" in lbl: sg+=amt
                elif "TDS" not in lbl and "ROUNDING" not in lbl and amt>0: tx+=amt
        save()

    if not rows: raise ValueError("No entries parsed. Use Detailed Purchase Register export from Tally.")
    return pd.DataFrame(rows)


def parse_gstr2b(fb):
    wb=openpyxl.load_workbook(io.BytesIO(fb),data_only=True)
    if "B2B" not in wb.sheetnames:
        raise ValueError("GSTR-2B must have a sheet named 'B2B'. Download Excel from portal.")
    ws=wb["B2B"]; rows=[]
    for row in ws.iter_rows(min_row=7,values_only=True):
        if not row[0] or not str(row[0]).strip(): continue
        rows.append({"GSTIN":clean_gstin(row[0]),"TradeNm":str(row[1] or "").strip(),
                     "InvNo":str(row[2] or "").strip(),"InvDate":row[4],
                     "Taxable":safe_float(row[8]),"IGST":safe_float(row[9]),
                     "CGST":safe_float(row[10]),"SGST":safe_float(row[11]),
                     "TotalTax":safe_float(row[9])+safe_float(row[10])+safe_float(row[11]),
                     "RCM":str(row[7] or "").upper()=="YES",
                     "ITCAvail":str(row[15] or "").strip()})
    if not rows: raise ValueError("No data in GSTR-2B B2B sheet.")
    return pd.DataFrame(rows)


def run_itc_reco(pr_bytes, b2b_bytes):
    df_t =parse_purchase_register(pr_bytes)
    df_2b=parse_gstr2b(b2b_bytes)

    def tkey(r): return r["GSTIN"] if r["GSTIN"] else re.sub(r"[^a-z0-9]","",r["Party"].lower())
    df_t["_key"] =df_t.apply(tkey,axis=1)
    df_2b["_key"]=df_2b.apply(lambda r: r["GSTIN"] if r["GSTIN"] else re.sub(r"[^a-z0-9]","",r["TradeNm"].lower()),axis=1)

    ts=df_t.groupby("_key").agg(T_Party=("Party","first"),T_GSTIN=("GSTIN","first"),
        T_Inv=("VchNo","count"),T_Value=("Value","sum"),T_IGST=("IGST","sum"),
        T_CGST=("CGST","sum"),T_SGST=("SGST","sum"),T_Tax=("TotalTax","sum"),
        T_Cat=("Category","first")).reset_index()

    gs=df_2b.groupby("_key").agg(G_Party=("TradeNm","first"),G_GSTIN=("GSTIN","first"),
        G_Inv=("InvNo","count"),G_Taxable=("Taxable","sum"),G_IGST=("IGST","sum"),
        G_CGST=("CGST","sum"),G_SGST=("SGST","sum"),G_Tax=("TotalTax","sum"),
        G_RCM=("RCM","any")).reset_index()

    reco=pd.merge(ts,gs,on="_key",how="outer").fillna(0)

    def get_status(r):
        cat=str(r.get("T_Cat",""))
        if "Custom" in cat: return "Custom Duty / Import"
        if r.get("G_RCM",False): return "RCM — Verify Books"
        ti,gi=r["T_Inv"],r["G_Inv"]
        if ti==0 and gi>0: return "Only in 2B"
        if gi==0 and ti>0: return "Only in Books"
        if abs(r["T_IGST"]-r["G_IGST"])<1 and abs(r["T_CGST"]-r["G_CGST"])<1 and abs(r["T_SGST"]-r["G_SGST"])<1:
            return "Matched"
        return "Mismatch"

    def get_remark(r):
        st=r["Status"]
        if st=="Matched": return "✓ ITC claimable — books match 2B"
        if st=="RCM — Verify Books": return "RCM — ensure self-invoice raised and tax paid in GSTR-3B"
        if st=="Custom Duty / Import": return "Custom duty/Import — verify IMPG in 2B; ITC available if ICEGATE reflected"
        if st=="Only in 2B": return "Supplier filed but not booked in Tally — check if purchase entry pending"
        if st=="Only in Books":
            if "Import" in str(r.get("T_Cat","")) or "Custom" in str(r.get("T_Cat","")):
                return "Import/Custom — may reflect in 2B next month (ICEGATE delay)"
            return "Booked in Tally but supplier not filed GSTR-1 — ITC blocked"
        if st=="Mismatch":
            parts=[]
            di,dc,ds=r["T_IGST"]-r["G_IGST"],r["T_CGST"]-r["G_CGST"],r["T_SGST"]-r["G_SGST"]
            if abs(di)>1: parts.append(f"IGST diff ₹{di:,.0f}")
            if abs(dc)>1: parts.append(f"CGST diff ₹{dc:,.0f}")
            if abs(ds)>1: parts.append(f"SGST diff ₹{ds:,.0f}")
            inv_diff=int(r["T_Inv"]-r["G_Inv"])
            if inv_diff!=0: parts.append(f"Invoice count diff {inv_diff:+d}")
            hint="Check prior-month invoices booked this month" if r["T_Tax"]>r["G_Tax"] else "Check invoices pending in books"
            return ("; ".join(parts)+f" — {hint}") if parts else hint
        return ""

    reco["Status"]=reco.apply(get_status,axis=1)
    reco["Remark"]=reco.apply(get_remark,axis=1)
    reco["D_IGST"]=reco["T_IGST"]-reco["G_IGST"]
    reco["D_CGST"]=reco["T_CGST"]-reco["G_CGST"]
    reco["D_SGST"]=reco["T_SGST"]-reco["G_SGST"]
    reco["D_Tax"] =reco["T_Tax"] -reco["G_Tax"]
    order={"Matched":0,"Mismatch":1,"RCM — Verify Books":2,"Custom Duty / Import":3,"Only in 2B":4,"Only in Books":5}
    reco["_s"]=reco["Status"].map(order).fillna(9)
    return reco.sort_values(["_s","T_Party"]).drop(columns=["_s"]).reset_index(drop=True),df_t,df_2b


# ═══════════════════════════════════════════════════════════════════════════════
#  SALES RECO
# ═══════════════════════════════════════════════════════════════════════════════
TALLY_SH=["Sales 17-18 Register","GST SALES (Local)","IGST (Sales)","Handling charges",
          "Storage Charges ","SALES TO SEZ (IGST)","STORAGE CHARGES - SEZ UNIT",
          "Sales Exempt","Detention Chgs","SHORTAGE IN TRANSIT","Freight Charges (Income)"]
EXEMPT_CATS={"Sales Exempt","Detention Chgs","SHORTAGE IN TRANSIT"}
_INV =["voucher no.","voucher no","invoice no","bill no","doc no"]
_PTY =["particulars","party","customer","party name"]
_DAT =["date","invoice date","bill date"]
_VAL =["value","taxable value","taxable","basic","amount"]
_GST =["gstin/uin","gstin"]

def _fc(hdrs,keys):
    hl=[str(h or "").strip().lower() for h in hdrs]
    for k in keys:
        for i,h in enumerate(hl):
            if k==h or (len(k)>4 and k in h): return i
    return None

def _smart_v(hdrs):
    hl=[str(h or "").lower() for h in hdrs]
    for i,h in enumerate(hl):
        if i>5 and any(kw in h for kw in ["storage charges","handling charges","igst sales",
            "gst sales","sales to sez","shortage","detention","freight charges"]): return i
    for k in _VAL:
        for i,h in enumerate(hl):
            if k==h or k in h: return i
    return None

def parse_sales_books(fb):
    wb=openpyxl.load_workbook(io.BytesIO(fb),data_only=True); rows=[]
    found_tally=any(s in wb.sheetnames for s in TALLY_SH)
    sheets=[s for s in TALLY_SH if s in wb.sheetnames] if found_tally else wb.sheetnames
    for sname in sheets:
        ws=wb[sname]; hr=None
        for i,row in enumerate(ws.iter_rows(min_row=1,max_row=20,values_only=True),1):
            rl=[str(v or "").lower() for v in row]
            if any(k in " ".join(rl) for k in ["voucher no","invoice no","bill no"]):
                hr=i; break
        if not hr: continue
        hdrs=[ws.cell(hr,c).value for c in range(1,ws.max_column+1)]
        inv_ci=_fc(hdrs,_INV); pty_ci=_fc(hdrs,_PTY); dat_ci=_fc(hdrs,_DAT)
        val_ci=_smart_v(hdrs);  gst_ci=_fc(hdrs,_GST)
        if inv_ci is None: continue
        for row in ws.iter_rows(min_row=hr+1,max_row=ws.max_row,values_only=True):
            inv=str(row[inv_ci] or "").strip()
            if not inv: continue
            p=str(row[pty_ci] or "").strip() if pty_ci else ""
            if p.lower() in ("total","grand total","(cancelled )","cancelled"): continue
            try: val=float(row[val_ci] or 0) if val_ci is not None else 0
            except: val=0
            rows.append({"Date":row[dat_ci] if dat_ci else row[0],
                         "Party":p,"GSTIN":clean_gstin(row[gst_ci]) if gst_ci else "",
                         "InvNo":inv,"Taxable":val,"Category":sname.strip()})
    if not rows: raise ValueError("No data found. File needs Invoice No + Value columns.")
    df=pd.DataFrame(rows)
    df["inv_norm"]=df["InvNo"].apply(lambda s: re.sub(r"[-\s]","",str(s).strip().upper()))
    return df.groupby("inv_norm").agg(Date=("Date","first"),Party=("Party","first"),
        GSTIN=("GSTIN","first"),InvNo=("InvNo","first"),
        Taxable=("Taxable","sum"),Category=("Category","first")).reset_index()

def parse_sales_portal(fb):
    wb=openpyxl.load_workbook(io.BytesIO(fb),data_only=True); rows=[]
    if "b2b, sez, de" not in wb.sheetnames:
        raise ValueError("Portal file needs sheet 'b2b, sez, de'. Download E-Invoice Excel from GST Portal.")
    ws=wb["b2b, sez, de"]
    for row in ws.iter_rows(min_row=5,max_row=ws.max_row,values_only=True):
        if not row[0] or not str(row[0]).strip(): continue
        rows.append({"GSTIN":clean_gstin(row[0]),"Party":str(row[1] or "").strip(),
                     "InvNo":str(row[2] or "").strip(),"Date":str(row[3] or ""),
                     "Taxable":safe_float(str(row[11] or "0").replace(",","")),
                     "IGST":safe_float(str(row[12] or "0").replace(",","")),
                     "CGST":safe_float(str(row[13] or "0").replace(",","")),
                     "SGST":safe_float(str(row[14] or "0").replace(",",""))})
    df=pd.DataFrame(rows)
    df["inv_norm"]=df["InvNo"].apply(lambda s: re.sub(r"[-\s]","",str(s).strip().upper()))
    return df

def run_sales_reco(portal_bytes,books_bytes):
    df_b=parse_sales_books(books_bytes); df_p=parse_sales_portal(portal_bytes)
    reco=pd.merge(df_b,df_p,on="inv_norm",how="outer",suffixes=("_b","_p"))
    reco["B_Tax"]=reco["Taxable_b"].fillna(0); reco["P_Tax"]=reco["Taxable_p"].fillna(0)
    reco["Diff"]=reco["B_Tax"]-reco["P_Tax"]
    def status(r):
        cat=str(r.get("Category",""))
        if cat in EXEMPT_CATS and r["P_Tax"]==0: return "Exempt / Bond Transfer"
        if r["B_Tax"]==0 and r["P_Tax"]==0: return "Skip"
        if r["B_Tax"]==0: return "Only in Portal"
        if r["P_Tax"]==0: return "Only in Books"
        if abs(r["Diff"])<1: return "Matched"
        return "Value Mismatch"
    reco["Status"]=reco.apply(status,axis=1)
    return reco[reco["Status"]!="Skip"].sort_values(["Status","inv_norm"]).reset_index(drop=True),df_b,df_p


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════
ITC_BG={"Matched":MATCH,"Mismatch":MIS,"RCM — Verify Books":RCM_C,
         "Custom Duty / Import":CUST_C,"Only in 2B":ONLY2,"Only in Books":ONLY1}
ITC_FC={"Matched":"1F6E1F","Mismatch":"C00000","RCM — Verify Books":"375623",
         "Custom Duty / Import":"843C0C","Only in 2B":"7F6000","Only in Books":"004080"}
SAL_BG={"Matched":MATCH,"Value Mismatch":MIS,"Only in Portal":ONLY2,"Only in Books":ONLY1,"Exempt / Bond Transfer":EXEMPT}
SAL_FC={"Matched":"1F6E1F","Value Mismatch":"C00000","Only in Portal":"7F6000","Only in Books":"004080","Exempt / Bond Transfer":"555555"}
SAL_RM={"Matched":"Books = Portal ✓",
         "Value Mismatch":"Same invoice, taxable value differs — verify",
         "Only in Portal":"Filed in GSTR-1 but not found in books — check other ledgers",
         "Only in Books":"In books but not in portal — check if filed or exempt supply",
         "Exempt / Bond Transfer":"Exempt / Bond Transfer — not e-invoiced (correct)"}

def _mhdr(ws,text,ncols):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    c=ws.cell(1,1); c.value=text
    c.font=Font(name="Arial",bold=True,size=12,color="FFFFFF")
    c.fill=F(TITLE); c.alignment=CTR(); ws.row_dimensions[1].height=28

def _detail_sheet(ws_new,df,cols,title,hc):
    ws_new.sheet_view.showGridLines=False; nc=len(cols)
    ws_new.merge_cells(f"A1:{get_column_letter(nc)}1"); h=ws_new.cell(1,1); h.value=title
    h.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); h.fill=F(hc); h.alignment=CTR()
    for ci,(hdr,w,fld) in enumerate(cols,1):
        c=ws_new.cell(2,ci); c.value=hdr; c.font=HF(9); c.fill=F(hc); c.alignment=CTR(); c.border=BD()
        ws_new.column_dimensions[get_column_letter(ci)].width=w
    for ri,row in df.iterrows():
        r=ri+3
        for ci,(hdr,w,fld) in enumerate(cols,1):
            c=ws_new.cell(r,ci); c.value=row.get(fld,""); c.border=BD(); c.font=DF(9)
            if fld=="Date": c.number_format="DD-MMM-YY"; c.alignment=CTR()
            elif fld in ("Value","IGST","CGST","SGST","TotalTax","Taxable","Taxable Value","IGST","CGST","SGST"):
                c.number_format=INR; c.alignment=RGT()
            else: c.alignment=LFT()
        ws_new.row_dimensions[r].height=14
    ws_new.freeze_panes="A3"

def build_itc_excel(reco,df_t,df_2b,client,period):
    wb=openpyxl.Workbook(); counts=reco["Status"].value_counts().to_dict()

    # Summary
    ws1=wb.active; ws1.title="Summary"; ws1.sheet_view.showGridLines=False
    _mhdr(ws1,f"{client}  —  GSTR-2B vs Tally ITC Reco  |  {period}",6)
    for ci,v in enumerate(["","Tally","GSTR-2B","Diff IGST","Diff CGST+SGST","Remark"],1):
        c=ws1.cell(3,ci); c.value=v; c.font=HF(9); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()
    srows=[("✅ Matched",counts.get("Matched",0),counts.get("Matched",0),0,0,"ITC fully claimable"),
           ("⚠️ Mismatch",counts.get("Mismatch",0),counts.get("Mismatch",0),"-","-","Investigate before claiming"),
           ("🔵 RCM",counts.get("RCM — Verify Books",0),0,0,0,"Self-invoice + pay in 3B"),
           ("🟠 Custom/Import",counts.get("Custom Duty / Import",0),counts.get("Custom Duty / Import",0),"-","-","Verify IMPG in 2B"),
           ("🟡 Only in 2B",0,counts.get("Only in 2B",0),"-","-","Check Tally booking"),
           ("⚪ Only in Books",counts.get("Only in Books",0),0,"-","-","Supplier not filed / Import delay")]
    for ri,rd in enumerate(srows,4):
        for ci,val in enumerate(rd,1):
            c=ws1.cell(ri,ci); c.value=val; c.border=BD(); c.font=DF(9,bold=(ci==1))
            c.alignment=RGT() if isinstance(val,(int,float)) else LFT()
        ws1.row_dimensions[ri].height=16
    # ITC totals
    ws1.merge_cells("A11:F11"); c=ws1.cell(11,1); c.value="ITC AMOUNTS (₹)"
    c.font=HF(9); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()
    for ci,v in enumerate(["","IGST","CGST","SGST","Total Tax",""],1):
        c=ws1.cell(12,ci); c.value=v; c.font=HF(8); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()
    irows=[("2B Total",df_2b["IGST"].sum(),df_2b["CGST"].sum(),df_2b["SGST"].sum(),df_2b["TotalTax"].sum()),
           ("Books Total",df_t["IGST"].sum(),df_t["CGST"].sum(),df_t["SGST"].sum(),df_t["TotalTax"].sum()),
           ("✅ Matched (Claimable)",
            reco[reco.Status=="Matched"]["G_IGST"].sum(),
            reco[reco.Status=="Matched"]["G_CGST"].sum(),
            reco[reco.Status=="Matched"]["G_SGST"].sum(),
            reco[reco.Status=="Matched"]["G_Tax"].sum())]
    for ri,rd in enumerate(irows,13):
        for ci,val in enumerate(rd,1):
            c=ws1.cell(ri,ci); c.value=val; c.border=BD(); c.font=DF(9,bold=(ri==15))
            if ci in (2,3,4,5): c.number_format=INR; c.alignment=RGT()
            else: c.alignment=LFT()
        ws1.row_dimensions[ri].height=16
    for col,w in zip("ABCDEF",[30,14,14,14,14,40]): ws1.column_dimensions[col].width=w

    # ITC Reco sheet
    ws2=wb.create_sheet("ITC Reco"); ws2.sheet_view.showGridLines=False
    _mhdr(ws2,f"{client}  —  ITC Reco (GSTR-2B vs Tally)  |  {period}",15)
    rcols=[("#",4),("Party Name",38),("GSTIN",20),
           ("Tally # Inv",9),("Tally IGST",15),("Tally CGST",15),("Tally SGST",15),("Tally Total Tax",15),
           ("2B # Inv",9),("2B IGST",15),("2B CGST",15),("2B SGST",15),("2B Total Tax",15),
           ("Status",20),("Remarks / Action",52)]
    for ci,(hdr,w) in enumerate(rcols,1):
        c=ws2.cell(3,ci); c.value=hdr; c.font=HF(8); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.row_dimensions[3].height=22

    for ri,row in reco.iterrows():
        r=ri+4; st=row["Status"]; bg=ITC_BG.get(st,"FFFFFF")
        party=str(row.get("T_Party") or row.get("G_Party") or "")
        gstin=str(row.get("T_GSTIN") or row.get("G_GSTIN") or "")
        vals=[ri+1,party,gstin,int(row["T_Inv"]),row["T_IGST"],row["T_CGST"],row["T_SGST"],row["T_Tax"],
              int(row["G_Inv"]),row["G_IGST"],row["G_CGST"],row["G_SGST"],row["G_Tax"],st,row["Remark"]]
        for ci,val in enumerate(vals,1):
            c=ws2.cell(r,ci); c.value=val; c.fill=F(bg); c.border=BD()
            if ci in (5,6,7,8,10,11,12,13): c.font=DF(9); c.alignment=RGT(); c.number_format=INR
            elif ci in (4,9): c.font=DF(9); c.alignment=CTR(); c.number_format=NUM
            elif ci==14: c.font=Font(name="Arial",bold=True,size=9,color=ITC_FC.get(st,"333")); c.alignment=CTR()
            else: c.font=DF(9); c.alignment=LFT()
        ws2.row_dimensions[r].height=15

    tr=len(reco)+4
    ws2.merge_cells(f"A{tr}:C{tr}"); tc=ws2.cell(tr,1); tc.value="TOTAL"
    tc.font=HF(9); tc.fill=F(NAVY); tc.alignment=CTR(); tc.border=BD()
    for ci in range(4,14):
        c=ws2.cell(tr,ci); c.font=HF(9); c.fill=F(NAVY); c.border=BD()
        if ci in (5,6,7,8,10,11,12,13):
            c.value=f"=SUM({get_column_letter(ci)}4:{get_column_letter(ci)}{tr-1})"; c.number_format=INR; c.alignment=RGT()
        elif ci in (4,9):
            c.value=f"=SUM({get_column_letter(ci)}4:{get_column_letter(ci)}{tr-1})"; c.number_format=NUM; c.alignment=CTR()
    ws2.row_dimensions[tr].height=17; ws2.freeze_panes="D4"

    _detail_sheet(wb.create_sheet("Tally Detail"),df_t,[
        ("Date",12,"Date"),("Party",36,"Party"),("GSTIN",20,"GSTIN"),
        ("Voucher No",20,"VchNo"),("Supplier Inv",20,"SupplierInv"),
        ("Value",15,"Value"),("IGST",14,"IGST"),("CGST",14,"CGST"),("SGST",14,"SGST"),("Category",22,"Category")
    ],f"Tally Purchase Register | {period}",BLUE)

    df_2b_o=df_2b.rename(columns={"TradeNm":"Trade Name","TotalTax":"Total Tax"})
    _detail_sheet(wb.create_sheet("GSTR-2B Detail"),df_2b_o,[
        ("GSTIN",22,"GSTIN"),("Trade Name",36,"Trade Name"),("Invoice No",20,"InvNo"),
        ("Taxable Value",15,"Taxable"),("IGST",14,"IGST"),("CGST",14,"CGST"),
        ("SGST",14,"SGST"),("Total Tax",14,"Total Tax"),("ITC Available",13,"ITCAvail"),("RCM",6,"RCM")
    ],f"GSTR-2B B2B | {period}",GREEN)

    # Legend
    wl=wb.create_sheet("Legend"); wl.sheet_view.showGridLines=False
    wl.merge_cells("A1:C1"); h=wl.cell(1,1); h.value="Color Legend"
    h.font=Font(name="Arial",bold=True,size=11,color="FFFFFF"); h.fill=F(TITLE); h.alignment=CTR()
    for i,(bg,st,desc) in enumerate([(MATCH,"Matched","ITC fully claimable"),(MIS,"Mismatch","Investigate before claiming"),
        (RCM_C,"RCM","Self-invoice + pay in GSTR-3B"),(CUST_C,"Custom Duty/Import","Verify IMPG in 2B"),
        (ONLY2,"Only in 2B","Check if purchase to be booked in Tally"),(ONLY1,"Only in Books","Supplier not filed or import pending")],3):
        wl.cell(i,1).value=st; wl.cell(i,1).fill=F(bg); wl.cell(i,1).font=DF(9,bold=True); wl.cell(i,1).border=BD()
        wl.merge_cells(f"B{i}:C{i}"); wl.cell(i,2).value=desc
        wl.cell(i,2).fill=F(bg); wl.cell(i,2).font=DF(9); wl.cell(i,2).border=BD(); wl.cell(i,2).alignment=LFT()
    wl.column_dimensions["A"].width=24; wl.column_dimensions["B"].width=55

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf


def build_sales_excel(reco,df_b,df_p,client,period):
    wb=openpyxl.Workbook(); counts=reco["Status"].value_counts().to_dict()
    b_tot=df_b["Taxable"].sum(); p_tot=df_p["Taxable"].sum()

    ws1=wb.active; ws1.title="Summary"; ws1.sheet_view.showGridLines=False
    _mhdr(ws1,f"{client}  —  GSTR-1 vs Books Sales Reco  |  {period}",5)
    for ci,v in enumerate(["","Books","Portal","Diff","Remark"],1):
        c=ws1.cell(3,ci); c.value=v; c.font=HF(9); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()
    for ri,rd in enumerate([("Total",len(df_b),len(df_p),len(df_b)-len(df_p),""),
        ("✅ Matched",counts.get("Matched",0),counts.get("Matched",0),0,"Clean"),
        ("⚠️ Mismatch",counts.get("Value Mismatch",0),counts.get("Value Mismatch",0),0,"Investigate"),
        ("🔵 Only Books",counts.get("Only in Books",0),0,counts.get("Only in Books",0),"Not filed"),
        ("🟡 Only Portal",0,counts.get("Only in Portal",0),-counts.get("Only in Portal",0),"Check books"),
        ("⚪ Exempt/Bond",counts.get("Exempt / Bond Transfer",0),0,"N/A","Not e-invoiced")],4):
        for ci,val in enumerate(rd,1):
            c=ws1.cell(ri,ci); c.value=val; c.border=BD(); c.font=DF(9,bold=(ci==1))
            if ci in (2,3,4) and isinstance(val,(int,float)): c.number_format="#,##0"; c.alignment=RGT()
            else: c.alignment=LFT()
        ws1.row_dimensions[ri].height=16
    for ci,v in enumerate(["","Books (₹)","Portal (₹)","Diff (₹)",""],1):
        c=ws1.cell(11,ci); c.value=v; c.font=HF(9); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()
    for ri,rd in enumerate([("Total Taxable",b_tot,p_tot,b_tot-p_tot),
        ("Matched",reco[reco.Status=="Matched"]["B_Tax"].sum(),reco[reco.Status=="Matched"]["P_Tax"].sum(),0),
        ("Exempt/Bond",reco[reco.Status=="Exempt / Bond Transfer"]["B_Tax"].sum(),0,reco[reco.Status=="Exempt / Bond Transfer"]["B_Tax"].sum()),
        ("Only Books",reco[reco.Status=="Only in Books"]["B_Tax"].sum(),0,reco[reco.Status=="Only in Books"]["B_Tax"].sum()),
        ("Only Portal",0,reco[reco.Status=="Only in Portal"]["P_Tax"].sum(),-reco[reco.Status=="Only in Portal"]["P_Tax"].sum())],12):
        for ci,val in enumerate(rd,1):
            c=ws1.cell(ri,ci); c.value=val; c.border=BD(); c.font=DF(9,bold=(ci==1))
            if ci in (2,3,4) and isinstance(val,(int,float)): c.number_format=INR; c.alignment=RGT()
            else: c.alignment=LFT()
        ws1.row_dimensions[ri].height=16
    for col,w in zip("ABCDE",[32,18,18,18,28]): ws1.column_dimensions[col].width=w

    ws2=wb.create_sheet("Invoice Reco"); ws2.sheet_view.showGridLines=False
    _mhdr(ws2,f"{client}  —  GSTR-1 vs Books Invoice Reco  |  {period}",11)
    rc=[("#",4),("Invoice No (Books)",22),("Invoice No (Portal)",22),("Date",12),
        ("Party",38),("Category",20),("Books (₹)",15),("Portal (₹)",15),("Diff (₹)",14),("Status",24),("Remarks",44)]
    for ci,(hdr,w) in enumerate(rc,1):
        c=ws2.cell(3,ci); c.value=hdr; c.font=HF(8); c.fill=F(NAVY); c.alignment=CTR(); c.border=BD()
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.row_dimensions[3].height=20
    for ri,row in reco.iterrows():
        r=ri+4; st=row["Status"]; bg=SAL_BG.get(st,"FFFFFF")
        b_inv=str(row.get("InvNo_b") or ""); p_inv=str(row.get("InvNo") or row.get("InvNo_p") or "")
        party=str(row.get("Party_b") or row.get("Party_p") or row.get("Party") or "")
        date=row.get("Date_b") or row.get("Date_p") or row.get("Date") or ""
        cat=str(row.get("Category",""))
        vals=[ri+1,b_inv,p_inv,date,party,cat,row["B_Tax"],row["P_Tax"],row["Diff"],st,SAL_RM.get(st,"")]
        for ci,val in enumerate(vals,1):
            c=ws2.cell(r,ci); c.value=val; c.fill=F(bg); c.border=BD()
            if ci in (7,8,9): c.font=DF(9); c.alignment=RGT(); c.number_format=INR
            elif ci==4: c.font=DF(9); c.number_format="DD-MMM-YY"; c.alignment=CTR()
            elif ci==1: c.font=DF(9); c.alignment=CTR()
            elif ci==10: c.font=Font(name="Arial",bold=True,size=9,color=SAL_FC.get(st,"333")); c.alignment=CTR()
            else: c.font=DF(9); c.alignment=LFT()
        ws2.row_dimensions[r].height=14
    ws2.freeze_panes="B4"

    df_b2=df_b.rename(columns={"InvNo":"Invoice No","Taxable":"Taxable Value"})
    df_p2=df_p.rename(columns={"InvNo":"Invoice No","Taxable":"Taxable Value"})
    _detail_sheet(wb.create_sheet("Books Detail"),df_b2,[
        ("Date",12,"Date"),("Party",40,"Party"),("Invoice No",22,"Invoice No"),
        ("GSTIN",22,"GSTIN"),("Category",22,"Category"),("Taxable Value",16,"Taxable Value")
    ],f"Books (Sales Register) | {period}",BLUE)
    _detail_sheet(wb.create_sheet("Portal Detail"),df_p2,[
        ("GSTIN",22,"GSTIN"),("Party",40,"Party"),("Invoice No",22,"Invoice No"),
        ("Date",12,"Date"),("Taxable Value",16,"Taxable Value"),
        ("IGST",14,"IGST"),("CGST",14,"CGST"),("SGST",14,"SGST")
    ],f"Portal — GSTR-1 E-Invoice | {period}",GREEN)

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf


# ═══════════════════════════════════════════════════════════════════════════════
#  UI
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown('<div class="hdr"><h1>📊 GST Audit Reconciliation Tool</h1><p>ITC Reco (GSTR-2B vs Tally) &nbsp;·&nbsp; Sales Reco (GSTR-1 vs Books) &nbsp;·&nbsp; Expert-level audit output</p></div>',unsafe_allow_html=True)

with st.sidebar:
    st.markdown("### Reco Type")
    reco_type=st.radio("",["🔵  ITC Reco  (2B vs Tally)","🟢  Sales Reco  (GSTR-1 vs Books)"],label_visibility="collapsed")
    st.markdown("---")
    client=st.text_input("Client Name",placeholder="Akin Chemicals Pvt Ltd")
    period=st.text_input("Period",      placeholder="January 2026")
    st.markdown("---")
    st.markdown("**Legend**")
    st.markdown("🟢 Matched · 🔴 Mismatch\n🟡 Only Portal/2B · 🔵 Only Books\n🟠 Custom/Import · ⚪ Exempt")

is_itc  ="ITC"   in reco_type
is_sales="Sales" in reco_type

if is_itc:
    st.markdown("### Upload Files — ITC Reco")
    c1,c2=st.columns(2)
    with c1:
        st.markdown("**File 1 — GSTR-2B Excel**")
        st.markdown('<div class="fmtbox">Portal → Returns → GSTR-2B → Download Excel (sheet: <code>B2B</code>)</div>',unsafe_allow_html=True)
        f1=st.file_uploader("GSTR-2B",type=["xlsx","xls"],key="i1",label_visibility="collapsed")
    with c2:
        st.markdown("**File 2 — Tally Purchase Register (Detailed)**")
        st.markdown('<div class="fmtbox">Gateway of Tally → Purchase Register → Detailed → Alt+E → Excel<br>Columns needed: Date · Particulars · GSTIN/UIN · Supplier Invoice No · Value · INPUT IGST · INPUT CGST · INPUT SGST</div>',unsafe_allow_html=True)
        f2=st.file_uploader("Purchase Register",type=["xlsx","xls"],key="i2",label_visibility="collapsed")
elif is_sales:
    st.markdown("### Upload Files — Sales Reco")
    c1,c2=st.columns(2)
    with c1:
        st.markdown("**File 1 — GSTR-1 E-Invoice Portal Excel**")
        st.markdown('<div class="fmtbox">GST Portal → E-Invoice → Download Excel<br>Must have sheet: <code>b2b, sez, de</code></div>',unsafe_allow_html=True)
        f1=st.file_uploader("GSTR-1 Portal",type=["xlsx","xls"],key="s1",label_visibility="collapsed")
    with c2:
        st.markdown("**File 2 — Tally Sales Register**")
        st.markdown('<div class="fmtbox">Any Tally sales register format accepted<br>Needs: Invoice No / Voucher No + Taxable Value columns</div>',unsafe_allow_html=True)
        f2=st.file_uploader("Sales Register",type=["xlsx","xls"],key="s2",label_visibility="collapsed")

st.markdown("---")
if f1 and f2:
    if not client.strip(): client="Client"
    if not period.strip(): period="Period"
    if st.button("▶  Run Reconciliation",type="primary",use_container_width=True):
        with st.spinner("Running..."):
            try:
                b1=f1.read(); b2=f2.read()
                if is_itc:
                    reco,df_t,df_2b=run_itc_reco(b1,b2)
                    cnt=reco["Status"].value_counts().to_dict()
                    buf=build_itc_excel(reco,df_t,df_2b,client,period)
                    fname=f"{client.replace(' ','_')}_ITC_Reco_{period.replace(' ','_')}.xlsx"
                    st.success("✅ ITC Reconciliation complete!")
                    cols=st.columns(6)
                    for col,lbl,key in zip(cols,["✅ Matched","⚠️ Mismatch","🟡 Only 2B","🔵 Only Books","🔵 RCM","🟠 Custom"],
                                           ["Matched","Mismatch","Only in 2B","Only in Books","RCM — Verify Books","Custom Duty / Import"]):
                        col.metric(lbl,cnt.get(key,0))
                    claimable=reco[reco.Status=="Matched"]["G_Tax"].sum()
                    st.info(f"**ITC Claimable (Matched): ₹{claimable:,.2f}**  ·  Total 2B: ₹{df_2b['TotalTax'].sum():,.2f}  ·  Total Books: ₹{df_t['TotalTax'].sum():,.2f}")
                    disp=reco[["T_Party","G_Party","T_IGST","G_IGST","T_CGST","G_CGST","D_Tax","Status","Remark"]].copy()
                    disp.columns=["Books Party","2B Party","Books IGST","2B IGST","Books CGST","2B CGST","Diff Tax","Status","Remark"]
                    st.dataframe(disp,use_container_width=True,height=380,
                        column_config={"Books IGST":st.column_config.NumberColumn(format="₹ %,.0f"),
                                       "2B IGST":st.column_config.NumberColumn(format="₹ %,.0f"),
                                       "Books CGST":st.column_config.NumberColumn(format="₹ %,.0f"),
                                       "2B CGST":st.column_config.NumberColumn(format="₹ %,.0f"),
                                       "Diff Tax":st.column_config.NumberColumn(format="₹ %,.0f")})
                elif is_sales:
                    reco,df_b,df_p=run_sales_reco(b1,b2)
                    cnt=reco["Status"].value_counts().to_dict()
                    buf=build_sales_excel(reco,df_b,df_p,client,period)
                    fname=f"{client.replace(' ','_')}_Sales_Reco_{period.replace(' ','_')}.xlsx"
                    st.success("✅ Sales Reconciliation complete!")
                    cols=st.columns(5)
                    for col,lbl,key in zip(cols,["✅ Matched","⚠️ Mismatch","🟡 Portal Only","🔵 Books Only","⚪ Exempt"],
                                           ["Matched","Value Mismatch","Only in Portal","Only in Books","Exempt / Bond Transfer"]):
                        col.metric(lbl,cnt.get(key,0))
                    disp=reco[["inv_norm","B_Tax","P_Tax","Diff","Status"]].copy()
                    disp.columns=["Invoice No","Books (₹)","Portal (₹)","Diff (₹)","Status"]
                    st.dataframe(disp,use_container_width=True,height=380,
                        column_config={"Books (₹)":st.column_config.NumberColumn(format="₹ %,.0f"),
                                       "Portal (₹)":st.column_config.NumberColumn(format="₹ %,.0f"),
                                       "Diff (₹)":st.column_config.NumberColumn(format="₹ %,.0f")})
                st.markdown("---")
                st.download_button("⬇️  Download Excel Reco File",data=buf,file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
            except Exception as e:
                st.error(f"❌ {e}")
                with st.expander("Details"): import traceback; st.code(traceback.format_exc())
else:
    st.info("👆 Upload both files above, then click Run.")
    with st.expander("📋 File format requirements"):
        st.markdown("""
**ITC Reco — Purchase Register (Tally Detailed format):**
One row per invoice. Required columns:
`Date` | `Particulars` | `GSTIN/UIN` | `Supplier Invoice No.` | `Value` | `INPUT IGST` | `INPUT CGST` | `INPUT SGST`

**Sales Reco — Portal (E-Invoice Excel):**
Must have sheet: `b2b, sez, de`

**Sales Reco — Books (Tally Sales Register):**
Any format. Needs: `Voucher No./Invoice No.` + `Value/Taxable Value` columns.
        """)
